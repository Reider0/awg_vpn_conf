import asyncpg
import os
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

class Database:
    def __init__(self):
        self.pool = None
        self.database_url = os.getenv("DATABASE_URL", "postgres://vpn:vpnpass@postgres:5432/vpndb")

    async def connect(self):
        for i in range(5):
            try:
                self.pool = await asyncpg.create_pool(dsn=self.database_url)
                await self.init_tables()
                print("Успешное подключение к БД")
                break
            except Exception as e:
                print(f"Попытка подключения к БД {i+1} неудачна: {e}")
                import asyncio
                await asyncio.sleep(2)

    async def init_tables(self):
        # Инициализация таблиц происходит через init.sql при первом запуске,
        # но здесь можно добавить проверки или миграции если нужно.
        pass

    async def execute(self, query, *args):
        if not self.pool: await self.connect()
        async with self.pool.acquire() as conn:
            return await conn.execute(query, *args)

    async def fetch_all(self, query, *args):
        if not self.pool: await self.connect()
        async with self.pool.acquire() as conn:
            return await conn.fetch(query, *args)
    
    async def fetch_val(self, query, *args):
        if not self.pool: await self.connect()
        async with self.pool.acquire() as conn:
            return await conn.fetchval(query, *args)

    async def device_set(self, uuid):
        query = "SELECT device FROM users WHERE uuid=$1"
        result = await self.fetch_all(query, uuid)
        if result and result[0]["device"]:
            return True
        return False

    async def get_all_users(self):
        query = "SELECT name, uuid, device, created_at, first_connected_at FROM users ORDER BY created_at DESC"
        return await self.fetch_all(query)
    
    async def get_user_by_uuid(self, uuid):
        query = "SELECT * FROM users WHERE uuid=$1"
        rows = await self.fetch_all(query, uuid)
        return rows[0] if rows else None

    # --- Работа с настройками (для Persistent Backup) ---
    async def set_setting(self, key, value):
        await self.execute(
            "INSERT INTO settings (key, value) VALUES ($1, $2) ON CONFLICT (key) DO UPDATE SET value = $2",
            key, str(value)
        )

    async def get_setting(self, key):
        return await self.fetch_val("SELECT value FROM settings WHERE key=$1", key)

    # --- Экспорт Excel с Картинками ---
    async def export_to_excel(self, path):
        users = await self.get_all_users()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "VPN Users"

        # Заголовки
        headers = ["Имя", "UUID", "Устройство", "Создано", "Вход", "Конфиг (Текст)", "QR Код"]
        ws.append(headers)

        # Настройка ширины колонок
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 50  # Текст конфига
        ws.column_dimensions['G'].width = 15  # QR код

        for index, u in enumerate(users, start=2): # start=2 т.к. 1 строка заголовки
            # Данные
            ws.cell(row=index, column=1, value=u['name'])
            ws.cell(row=index, column=2, value=u['uuid'])
            ws.cell(row=index, column=3, value=u['device'] or "")
            ws.cell(row=index, column=4, value=u['created_at'].strftime("%Y-%m-%d %H:%M") if u['created_at'] else "")
            ws.cell(row=index, column=5, value=u['first_connected_at'].strftime("%Y-%m-%d %H:%M") if u['first_connected_at'] else "")
            
            # Читаем текст конфига
            conf_path = f"/volumes/configs/{u['name']}.conf"
            conf_text = "Файл не найден"
            if os.path.exists(conf_path):
                with open(conf_path, "r") as f:
                    conf_text = f.read()
            
            # Вставляем текст конфига с переносом строк
            cell_conf = ws.cell(row=index, column=6, value=conf_text)
            cell_conf.alignment = Alignment(wrap_text=True)

            # Вставляем QR код
            qr_path = f"/volumes/configs/{u['name']}.png"
            if os.path.exists(qr_path):
                try:
                    img = ExcelImage(qr_path)
                    # Подгоняем размер под ячейку (примерно 100x100 px)
                    img.width = 100
                    img.height = 100
                    # Привязываем к ячейке G{index}
                    anchor = f"G{index}"
                    ws.add_image(img, anchor)
                    # Увеличиваем высоту строки, чтобы картинка влезла
                    ws.row_dimensions[index].height = 80
                except Exception as e:
                    print(f"Ошибка вставки QR для {u['name']}: {e}")

        wb.save(path)
        return path

db = Database()